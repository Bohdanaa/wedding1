from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, FileResponse
import requests
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from telegram import Bot, Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CallbackQueryHandler, CommandHandler,CallbackContext
import asyncio




app = FastAPI()
templates = Jinja2Templates(directory="webpage")
bot_token = "6139494128:AAFF81pUP18MzObbGas48aBnlQnxn_9C42U"
bot = Bot(token=bot_token)
updater = Updater(token=bot_token, use_context=True)
dispatcher = updater.dispatcher


# Налаштування статичних файлів
app.mount("/webpage", StaticFiles(directory="webpage"), name="static")

data=[]

def send_telegram_message(name: str, presence: str, drinks: list, car:list):
    message = f"Імя: {name}\nПрисутність: {presence}\nНапої: {', '.join(drinks)}\nДоїзд: {', '.join(car)}"

    bot_token = "6139494128:AAFF81pUP18MzObbGas48aBnlQnxn_9C42U"
    chat_id = "500842187"

    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    data = {"chat_id": chat_id, "text": message}
    #excel
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("Отримати Excel файл", callback_data="get_file")],
                                 [InlineKeyboardButton("Отримати відсортований Excel файл", callback_data="sort_file")],
                                 [InlineKeyboardButton("Отримати підрахунок",callback_data="calculator_file")]])

  
    #noexcel

    response = requests.post(url, json=data)

    bot.send_message(chat_id=chat_id, text=message, reply_markup=keyboard)
    return response.json()
def create_excel_file(data):
    try:
        wb = load_workbook('data.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
    headers = ['Ім`я', 'Напої','Доїзд']
    for row_data in data:
        row = [row_data['name'], row_data['presence']]
        drinks = ['Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю']
        row.append(', '.join(row_data['drinks']))
        row.append(', '.join(row_data['car']))
        sheet.append(row)
    for col_num, headers in enumerate(headers, start=1):
        col_letter = chr(65 + col_num)
        sheet[f'{col_letter}1'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 15

    

    wb.save('data.xlsx')

   
    print('Файл "data.xlsx" створено/оновлено успішно.')

# Create excel
def start(update: Update, context: CallbackContext):
    chat_id = update.effective_chat.id
    response_message = "Привіт! Я бот для отримання даних."
    context.bot.send_message(chat_id=chat_id, text=response_message)



def create_sorted_excel_file(data):
    wb = Workbook()
    sheet = wb.active
    headers = ['Ім`я', 'Присутність', 'Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю','Доїзд']
    sheet.append(headers)
    for row_data in sorted(data, key=lambda x: x['name']):
        row = [row_data['name'], row_data['presence']]

        drinks = ['Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю']
      
        for drink in drinks:
            if drink in row_data['drinks']:
                row.append('✔')
            else:
                row.append('✗')
        row.append(', '.join(row_data['car']))
        sheet.append(row)

    for col_num, headers in enumerate(headers, start=1):
        col_letter = chr(65 + col_num)
        sheet[f'{col_letter}1'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 15

    wb.save('sorted_data.xlsx')
    print('Файл "sorted_data.xlsx" створено/оновлено успішно.')
def create_calculator_excel_file(data):
    wb = Workbook()
    sheet = wb.active
    headers = ['Ім`я', 'Присутність', 'Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю','На машині', 'Без машини']
    sheet.append(headers)
    for row_data in sorted(data, key=lambda x: x['name']):
        row = [row_data['name'], row_data['presence']]

        drinks = ['Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю']
        cars=['Приїду на машині', 'Без машини']
        for drink in drinks:
            if drink in row_data['drinks']:
                row.append('1')
            else:
                row.append('0')
        for car in cars:
            if car in row_data['car']:
                row.append('1')
            else:
                row.append('0')
        sheet.append(row)

    for col_num, headers in enumerate(headers, start=1):
        col_letter = chr(65 + col_num)
        sheet[f'{col_letter}1'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 15
    global total_names, total_wines_red, total_wines_white, total_beers, total_rom, total_viski, total_no, total_vodka, total_samogon
    total_names = len(data)  # Загальна кількість учасників
    total_presence = sum(1 for row_data in data if row_data['presence'] == 'Присутній')
    total_absence = sum(1 for row_data in data if  row_data['presence'] == 'Відсутній')
    total_50 = sum(1 for row_data in data if row_data['presence'] == 'Пізніше')
    total_wines_red = sum(1 for row_data in data if 'Вино червоне' in row_data['drinks'])
    total_wines_white = sum(1 for row_data in data if 'Вино біле' in row_data['drinks'])
    total_beers = sum(1 for row_data in data if 'Пиво' in row_data['drinks'])
    total_vodka = sum(1 for row_data in data if 'Горілка' in row_data['drinks'])
    total_samogon = sum(1 for row_data in data if 'Самогонка' in row_data['drinks'])
    total_viski = sum(1 for row_data in data if 'Віскі' in row_data['drinks'])
    total_rom = sum(1 for row_data in data if 'Ром' in row_data['drinks'])
    total_no = sum(1 for row_data in data if 'Не вживаю' in row_data['drinks'])
    total_cars = sum(1 for row_data in data if 'Приїду на машині' in row_data['car'])
    total_no_cars = sum(1 for row_data in data if 'Без машини' in row_data['car'])
    
    total_presence = 0
    total_absence = 0
    total_50 = 0
    for row_data in data:
        if row_data['presence'] == 'Присутній':
            total_presence+=1
           
        elif row_data['presence'] == 'Відсутній':
            total_absence+=1
           
        elif row_data['presence'] == 'Пізніше':
            total_50+=1
        selected_total=total_50+total_presence+total_absence
            

    print(selected_total)
    print(type(total_wines_red))
    total_row = ['Підсумок',selected_total,total_wines_red, total_wines_white, total_beers, total_vodka, total_samogon, total_viski, total_rom, total_no, total_cars,total_no_cars]
    sheet.append(total_row)
    total_row_labels = ['Будуть', 'Не будуть', 'Не знаю', 'Відмітилось']
    total_row_values = [ total_presence, total_absence,total_50, selected_total]
    sheet.append(total_row_labels)
    sheet.append(total_row_values)
   


    wb.save('calculator_data.xlsx')
    print('Файл "calculator_data.xlsx" створено/оновлено успішно.')
#Видалення даних
def clear_excel_data():
    try:
        wb = load_workbook('data.xlsx')
        sheet = wb.active
        sheet.delete_rows(2, sheet.max_row)  # Видалити рядки від другого рядка до останнього
        wb.save('data.xlsx')
        print('Дані успішно видалено з таблиці.')
    except FileNotFoundError:
        print('Файл "data.xlsx" не знайдено.')



@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    with open("webpage/index.html", "r") as file:
        html_content = file.read()
    return HTMLResponse(content=html_content, status_code=200)


@app.post("/submit")
async def submit(request: Request):
   
    form = await request.form()
    name = form.get("name")
    presence = form.get("presence")
    drinks = form.getlist("drinks")
    car = form.getlist("car")
    await asyncio.sleep(1) 
    data.append({'name': name, 'presence': presence, 'drinks': drinks, 'car': car})
    create_excel_file(data)
    create_sorted_excel_file(data)
    create_calculator_excel_file(data)
    

    send_telegram_message(name, presence, drinks,car)
    

    return templates.TemplateResponse("index.html", {"request": request, "message": "Повідомлення успішно відправлено"})


@app.get("/get-file")
def get_file():
    return FileResponse("data.xlsx", filename="data.xlsx", media_type="application/octet-stream")


@app.get("/get-sorted-file")
def get_sorted_file():
    return FileResponse("sorted_data.xlsx", filename="sorted_data.xlsx", media_type="application/octet-stream")

@app.get("/get-sorted-file")
def get_calculator_file():
    return FileResponse("calculator_data.xlsx", filename="calculator_data.xlsx", media_type="application/octet-stream")

def handle_button_click(update: Update, context: CallbackContext):
    chat_id = update.callback_query.message.chat_id
    bot.send_document(chat_id=chat_id, document=open('data.xlsx', 'rb'))

def handle_sorted_button_click(update: Update, context: CallbackContext):
    chat_id = update.callback_query.message.chat_id
    bot.send_document(chat_id=chat_id, document=open('sorted_data.xlsx', 'rb'))

def handle_calculator_button_click(update: Update, context: CallbackContext):
    chat_id = update.callback_query.message.chat_id
    bot.send_document(chat_id=chat_id, document=open('calculator_data.xlsx', 'rb'))




start_handler = CommandHandler('start', start)
dispatcher.add_handler(start_handler)

updater.start_polling()
dispatcher.add_handler(CallbackQueryHandler(handle_button_click, pattern='get_file'))
dispatcher.add_handler(CallbackQueryHandler(handle_sorted_button_click, pattern='sort_file'))
dispatcher.add_handler(CallbackQueryHandler(handle_calculator_button_click,pattern='calculator_file'))
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
   
   
